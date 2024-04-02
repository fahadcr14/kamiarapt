
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook

from selenium.webdriver.common.keys import Keys
#from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import requests
from bs4 import BeautifulSoup
import random
import os
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import re
import mysql.connector

prefix_http = "https://shop.cardinalhealth.ca"
header = ['Name', 'Description', 'Catalogue_Number', 'Price', 'Availability', 'Manufacturer_Numbe',
          'Invoice_Description', 'Feature_Property', 'Sterility', 'Link', 'Brand_Name', 'Outer_Diameter',
          'Needle_Length', 'Volume', 'color', 'Inner_Diameter', 'size', 'age', 'Shape', 'Catheter_Length',
          'Composition_Ingredient']



def create_chrome_drive():
    print("begin create chrome drive")
    #driver_path = ChromeDriverManager().install()
    #service = Service(driver_path)
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver


def login():
    global driver
    print("begin login")
    time.sleep(5)
    #driver = create_chrome_drive()
    driver.get(
        "https://shop.cardinalhealth.ca/en/chc/cardinal-health#facet:&productBeginIndex:0&orderBy:&pageView:list&minPrice:&maxPrice:&pageSize:&")

    # wait button
    button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="button-header-sign-in"]'))
    )
    # click button
    button.click()

    # driver.find_element("xpath", "//button[@data-testid='button-header-sign-in']").click()
    sleep(10)

    driver.find_element("xpath", "//input[@name='logonId']").send_keys('server_crawler')
    driver.find_element("xpath", "//input[@name='logonPassword']").send_keys('0RTttPJP3SB%')
    sleep(1)
    driver.find_element("xpath", "//button[@data-testid='button-sign-in-submit']").click()
    sleep(5)
    print(f'Login Successfully')
    return driver


# get category url
def main_link():
    global driver
    #driver = create_chrome_drive()
    link = 'https://shop.cardinalhealth.ca/en/chc/cardinal-health#facet:&productBeginIndex:0&orderBy:&pageView:list&minPrice:&maxPrice:&pageSize:&'
    driver.get(link)
    time.sleep(5)
    button = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, f'//*[@id="root"]/div[1]/div/div[1]/header/div[2]/div/div/div[1]/div/button'))
    )
    print("Button element is present on the page.")
    #  wait a moment，make sure the page load
    #driver.implicitly_wait(10)  #
    #time.sleep(10)
    # click button
    button.click()
    #button.click()

    driver.implicitly_wait(10)
    time.sleep(10)
    link_href = []

    for index in range(1,11):
        print(index)
        # move to element
        #target_element = driver.find_element(By.XPATH, f'//*[@id="root"]/div[1]/div/div[1]/header/div[2]/div/div/div[1]/div/button')
        target_element = driver.find_element(By.XPATH,
                                                 f"//*[@id='MENU_POPPER_allCategories']/div/div/div/div[1]/div/div[{index}]/span/p")

        # use ActionChains move mouth on element
        actions = ActionChains(driver)
        actions.move_to_element(target_element).perform()

        print("Element is present on the page.")
        dynamic_content = driver.page_source
        # print(dynamic_content)
        #
        # use Beautiful Soup parse dynamic Web page
        soup = BeautifulSoup(dynamic_content, 'html5lib')

        # print('soup:', soup)
        Link_web_element = soup.find('div', {'id': 'MENU_POPPER_allCategories'}).find_all('a')
        print("Link_web_element:", Link_web_element)
        for link_web in Link_web_element:
            Link = link_web['href']
            link_href.append(prefix_http + Link)

    with open('Main_Category_Link.txt', 'w') as f:
        f.write('')
    # try:
    # Link_web_element = soup.find('div', {'id': 'MENU_POPPER_allCategories'}).find_all('a')
    # print("Link_web_element:", Link_web_element)
    # for link_web in Link_web_element:
    print(f'done main link')
    for link_web in link_href:
        # Link = link_web['href']
        # print(Link)
        try:
            with open('Main_Category_Link.txt', 'a') as f:
                # f.write(prefix_http + Link)
                f.write(link_web)
                f.write('\n')
        except Exception as e:
            return False
    # except Exception:
    #     pass
    print(f'going main link')

    return driver
    #driver.close()


def product_link():
    # check file exist
    if os.path.exists('Product_link.txt'):
        os.remove('Product_link.txt')

    print("begin product link!")
    # driver = login()
    #driver = create_chrome_drive()

    with open(r'Main_Category_Link.txt') as f:
        lines = f.readlines()
        for link in lines:
            link = link.replace('\n', '')
            print("---------------------------------------------------------------")
            print(link)

            driver.get(link)
            x = 0
            total_page = 1
            try:
                WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div.product-listing-container.productListingWidget.top-margin-3")))
                # driver.implicitly_wait(10)  #
                time.sleep(10)

                print("Element is present on the page.")
            

                total_page = driver.find_element("xpath", "//nav[@aria-label='pagination navigation']")
                # print(total_page.text)

                total_page = total_page.text[-1]
                total_page = int(total_page)
                print(total_page)
            except Exception:
                pass

            # x = total_page
            while True:
                print("begin get product urls!")
                try:

                    #product_links = driver.find_elements('xpath', '//div[@class="sc-lizKOf bbWZbU"]/a')
                    #2.0 faddy
                    product_links = driver.find_elements(By.CLASS_NAME, 'sc-dAlyuH.eoUdzI.sc-eoVZPG.ihtLNB.product-card')

                    Product_page_ = [link.get_attribute('href') for link in product_links]
                    print(f'Product pages {Product_page_}')
                    for product_link in Product_page_:
                        print(product_link)

                        with open('Product_link.txt', 'a+') as f:
                            print(f'added')
                            f.write(product_link)
                            f.write('\n')

                except Exception:
                    break

                if x == total_page:
                    break

                x += 1

                try:
                    next_page_button = driver.find_element(By.CSS_SELECTOR,
                                                           "button.MuiPaginationItem-root[aria-label='Go to next page']")
                    next_page_button.click()
                    sleep(10)
                except Exception:
                    print("----------------------- Not any page ---------------------------")
                    # traceback.print_exc()
                    break

            print("---------------------------------------------------------------")


# header = ['Name', 'Description', 'Catalogue_Number', 'Price', 'Availability', 'Manufacturer_Numbe',
#           'Invoice_Description', 'Feature_Property', 'Sterility', 'Link', 'Brand_Name', 'Outer_Diameter',
#           'Needle_Length', 'Volume', 'color', 'Inner_Diameter', 'size', 'age', 'Shape', 'Catheter_Length',
#           'Composition_Ingredient']

def information():
    global driver
    save_file_name = "data_res.xlsx"
    wb = load_workbook(filename=save_file_name)
    sheet = wb['Sheet1']
    counter = 2

    # driver = create_chrome_drive()
    #river = login()
    # create database connection
    #db = create_mysql_connection()
    # delete record in table
    #print("-------------------------- Record deleted Begin------------------------------------")
    #mydata = #db.cursor()
    #mydata.execute("TRUNCATE TABLE information")
    #db.commit()
    #print("-------------------------- Record deleted Finish------------------------------------")

    unique_set = set()
    with open('Product_link.txt') as f:
        lines = f.readlines()
        for Product_link in lines:
            Product_link = Product_link.replace('\n', '')

            unique_set.add(Product_link)

    begin = 0
    for Product_link in unique_set:

        print("---------------------------------------------------------------")
        print(Product_link)
        print("---------------------------------------------------------------")
        soup = BeautifulSoup()
        try:
            # wait
            driver.get(Product_link)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, "MuiTabs-root"))
            )

            # dynamic_content = driver.page_source
            # soup = BeautifulSoup(dynamic_content, 'html5lib')
            #
            sleep(2)
        except Exception:
            pass

        # try:
        #     user_name = driver.find_element('xpath',
        #                                     '//p[@class="MuiTypography-root sc-dcJsrY bXFCSC MuiTypography-body1"]').text
        #     if 'crawler server' not in user_name:
        #         driver = login()
        # except Exception:
        #     pass
        #
        #         # sheet.cell(row=counter , column=11).value=Product_link
        #

        name = ""
        try:
            name = driver.find_element('xpath', '//h4[@itemprop="name"]').text
            print("name:", name)
            # sheet.cell(row=counter, column=1).value = name
        except Exception:
            name = ''
            pass

        brand = ""
        try:
            brand = driver.find_element('xpath',
                                        '/html/body/div[1]/div[1]/div/div[2]/div/div/div[2]/div[1]/div/div/div[1]/a').text

            print("brand:", brand)
            # sheet.cell(row=counter, column=12).value = brand
        except Exception:
            brand = ''
            pass

        picture_url = ""
        try:
            picture_url = driver.find_element(By.XPATH, '//img[@itemprop="image"]').get_attribute('src')
            # picture_url = driver.find_element('xpath',
            #                                   '/html/body/div[1]/div[1]/div/div[2]/div/div/div[2]/div[1]/div/div/div[1]/div/div/img') \
            #     .get_attribute('src')
            print("picture_url:", picture_url)
            # sheet.cell(row=counter, column=22).value = picture_url
        except Exception:
            picture_url = ''
            pass

        # if begin == 0:
        #     begin += 1
        #     driver.get(Product_link)
        #     sleep(10)

        sleep(5)
        table_list = []
        try:

            dynamic_content = driver.page_source
            # print(dynamic_content)
            soup = BeautifulSoup(dynamic_content, 'html5lib')
            # find table
            table = soup.find('table', {'data-testid': 'table'})
            # Get header
            headers = [th.text.strip() for th in table.find('thead').find('tr').find_all('th')]
            print(headers)

            # Get table data rows
            rows = table.find('tbody').find_all('tr')

            result_dict = {}
            for i in range(0, len(rows)):
                print(rows[i])
                if i % 2 == 0:
                    tds = rows[i].find_all('td')
                    for j in range(0, len(tds)):
                        print(j, tds[j])
                        if j == 0:
                            if tds[0].find("div").text == "SKU":
                                result_dict['SKU'] = tds[0].find("span").text
                                print(tds[0].find("div").text)
                                print(tds[0].find("span").text)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Price":
                            price_divs = tds[j].find_all('div')

                            for price_div in price_divs:
                                print("price:", price_div)
                                if is_price(price_div.text):
                                    print("price.text:", price_div.text)
                                    result_dict['Price'] = price_div.text
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Availability":
                            if tds[j].find('span') is not None:
                                Availability = tds[j].find('span').text
                                result_dict['Availability'] = Availability
                                print("Availability:", Availability)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Shape":
                            original_string = tds[j].text
                            Shape = original_string.replace("Shape", "")
                            result_dict['Shape'] = Shape
                            print("shape:", Shape)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Age":
                            original_string = tds[j].text
                            Age = original_string.replace("Age", "")
                            result_dict['Age'] = Age
                            print("Age:", Age)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Size":
                            original_string = tds[j].text
                            Size = original_string.replace("Size", "")
                            result_dict['Size'] = Size
                            print("Size:", Size)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Outer Diameter":
                            original_string = tds[j].text
                            Outer_Diameter = original_string.replace("Outer Diameter", "")
                            result_dict['Outer_Diameter'] = Outer_Diameter
                            print("Outer_Diameter:", Outer_Diameter)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Inner Diameter":
                            original_string = tds[j].text
                            Inner_Diameter = original_string.replace("Inner Diameter", "")
                            result_dict['Inner_Diameter'] = Inner_Diameter
                            print("Inner_Diameter:", Inner_Diameter)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Composition / Ingredient":
                            original_string = tds[j].text
                            Composition_Ingredient_Diameter = original_string.replace("Composition / Ingredient", "")
                            result_dict['Composition_Ingredient_Diameter'] = Composition_Ingredient_Diameter
                            print("Composition_Ingredient_Diameter:", Composition_Ingredient_Diameter)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Needle Length":
                            original_string = tds[j].text
                            Needle_Length = original_string.replace("Needle Length", "")
                            result_dict['Needle_Length'] = Needle_Length
                            print("Needle_Length:", Needle_Length)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Length":
                            original_string = tds[j].text
                            Length = original_string.replace("Length", "")
                            result_dict['Length'] = Length
                            print("Length:", Length)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Colour":
                            original_string = tds[j].text
                            Colour = original_string.replace("Colour", "")
                            result_dict['Colour'] = Colour
                            print("Colour:", Colour)
                        elif tds[j].find("div") is not None and tds[j].find("div").text == "Width":
                            original_string = tds[j].text
                            Width = original_string.replace("Width", "")
                            result_dict['Width'] = Width
                            print("Width:", Width)

                if i % 2 == 1:
                    texts = [p.text.strip() for p in rows[i].find_all('p')]
                    print("lens:", len(texts))
                    print(texts)
                    for i in range(0, len(texts), 2):
                        key = texts[i]
                        value = texts[i + 1]
                        result_dict[key] = value

                    print(result_dict)
                    table_list.append(result_dict)
                    result_dict = {}

        except Exception as e:
            # Catch exceptions and print error messages
            print(f"发生了异常：{e}")

        Description = ""
        try:
            # Find the parent element that contains all buttons
            tabs_container = driver.find_element(By.CLASS_NAME, "MuiTabs-flexContainer")

            # Find all buttons
            buttons = tabs_container.find_elements(By.TAG_NAME, "button")

            # Click each button in turn
            for button in buttons:
                button_text = button.find_element(By.CSS_SELECTOR, '.MuiTab-wrapper').text
                # print("button's name：", button_text)
                button.click()
                # if button_text == "Features":
                #     p_elements = driver.find_elements(By.CSS_SELECTOR, 'div[role="tabpanel"] p')
                #     result_string = '\n'.join([p.text for p in p_elements])
                #     print("Features：", result_string)
                # sheet.cell(row=counter, column=2).value = result_string
                if button_text == "Description":
                    p_elements = driver.find_elements(By.CSS_SELECTOR, 'div[role="tabpanel"] p')
                    Description = '\n'.join([p.text for p in p_elements])
                    print("Description：", Description)
                    # sheet.cell(row=counter, column=2).value = Description
                    break
                sleep(1)
        except Exception:
            pass

        # Create an empty list to store the data to be inserted
        data_to_insert = []

        if len(table_list) == 0:
            sheet.cell(row=counter, column=1).value = name
            sheet.cell(row=counter, column=2).value = Description
            sheet.cell(row=counter, column=11).value = brand
            sheet.cell(row=counter, column=22).value = picture_url
            sheet.cell(row=counter, column=10).value = Product_link
            counter += 1

            # Add data to list in bulk insert
            data_to_insert.append((name, Description, "", "", "",
                                   "", "", "", "", Product_link,
                                   brand, "", "", "", ""
                                                      "", "", "", "", "", "", picture_url))


        else:

            for i in range(len(table_list)):
                my_table = table_list[i]
                for key, value in my_table.items():
                    if key == "Packaging": sheet.cell(row=counter, column=9).value = value
                    if key == "Feature / Property": sheet.cell(row=counter, column=8).value = value
                    if key == "Invoice Description": sheet.cell(row=counter, column=7).value = value
                    if key == "MPN": sheet.cell(row=counter, column=6).value = value
                    if key == "Volume": sheet.cell(row=counter, column=14).value = value
                    if key == "SKU": sheet.cell(row=counter, column=3).value = value
                    if key == "Price": sheet.cell(row=counter, column=4).value = value
                    if key == "Availability": sheet.cell(row=counter, column=5).value = value
                    if key == "Shape": sheet.cell(row=counter, column=19).value = value
                    if key == "Age": sheet.cell(row=counter, column=18).value = value
                    if key == "Size": sheet.cell(row=counter, column=17).value = value
                    if key == "Outer_Diameter": sheet.cell(row=counter, column=12).value = value
                    if key == "Inner_Diameter": sheet.cell(row=counter, column=16).value = value
                    if key == "Composition_Ingredient_Diameter": sheet.cell(row=counter, column=21).value = value
                    if key == "Needle_Length": sheet.cell(row=counter, column=13).value = value
                    if key == "Colour": sheet.cell(row=counter, column=15).value = value
                    sheet.cell(row=counter, column=1).value = name
                    sheet.cell(row=counter, column=2).value = Description
                    sheet.cell(row=counter, column=11).value = brand
                    sheet.cell(row=counter, column=22).value = picture_url
                    sheet.cell(row=counter, column=10).value = Product_link

                    wb.save(save_file_name)
                    print(f'Saved -----------------------------------------------')
                    # Add data to list in bulk insert
                    data_to_insert.append((name, Description, my_table.get("SKU", ""), my_table.get("Price", ""),
                                           my_table.get("Availability", ""),
                                           my_table.get("MPN", ""), my_table.get("Invoice Description", ""),
                                           my_table.get("Feature / Property", ""), my_table.get("Packaging", ""),
                                           Product_link,
                                           brand, my_table.get("Outer_Diameter", ""), my_table.get("Needle_Length", ""),
                                           my_table.get("Volume", ""), my_table.get("Colour", ""),
                                           my_table.get("Inner_Diameter", ""), my_table.get("Size", ""),
                                           my_table.get("Age", ""), my_table.get("Shape", ""), "",
                                           my_table.get("Composition_Ingredient_Diameter", ""), picture_url))
                counter += 1

        


        

    wb.save(save_file_name)


def is_price(s):
    pattern = r'^[¥$€£]?[0-9,]+(?:\.[0-9]+)?$'  # Match numbers, can have currency symbols, commas and decimal points
    return re.match(pattern, s) is not None






if __name__ == '__main__':
    driver=create_chrome_drive()

    login()
    main_link()

    product_link()
    information()
    driver.quit()





